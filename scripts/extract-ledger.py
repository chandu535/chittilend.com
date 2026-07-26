#!/usr/bin/env python3
"""
Extracts the `working` sheet of the legacy ledger into reviewable JSON, plus the
embedded photos.

    python3 scripts/extract-ledger.py /path/to/link.xlsx

Writes to .extract/ (gitignored — the output contains borrower names, phone numbers
and photographs, and must never enter the repo).

Nothing is written to any database. Seeding is a separate, later step so the mapping
can be checked first.

Layout of the sheet
    A  S No          loan serial
    B  Date          date the loan was given
    C  Name          borrower name (repeats — rows are loans, not borrowers)
    D  Phone         borrower identity key
    E  Amount(OUT)   cash handed over = principal x 0.99  (present on ~40% of rows)
    F  Amount(IN)    total repayable  = principal x 1.25  (present on ~99%)
    G..AI            29 monthly payment cells, Mar 2024 .. Jul 2026
    AJ Photo, AK Aadhar   images, anchored by row in xl/drawings
    AL Location      Google Maps hyperlink
    AM AREA, AN Surity

The IN/OUT ratio is 1.263 on 156 of 169 rows, which is exactly 1.25 / 0.99 — the
1% service charge and 25% markup the app already models.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python3 -m pip install --user openpyxl")

OUT_DIR = Path('.extract')
PHOTO_DIR = OUT_DIR / 'photos'

SHEET = 'working'
COL_SNO, COL_DATE, COL_NAME, COL_PHONE, COL_OUT, COL_IN = 1, 2, 3, 4, 5, 6
MONTH_COL_FIRST, MONTH_COL_LAST = 7, 35          # G .. AI
COL_PHOTO, COL_AADHAAR, COL_LOCATION, COL_AREA, COL_SURETY = 36, 37, 38, 39, 40

# Malformed dates confirmed by hand — the only 8 of 429 rows no parser could resolve
# safely. Keyed on the raw cell text exactly as it appears in the sheet.
DATE_CORRECTIONS = {
    '15-07-20205': '15-07-2025',
    '22-07-20205': '22-07-2025',
    '05-10-0204':  '05-10-2024',
    '21- 04-2025': '21-04-2025',
    '11.-06-2025': '11-06-2025',
    '02-07.-2025': '02-07-2025',
}

# Confirmed by hand, keyed on sheet row rather than cell text so the override cannot
# accidentally match another row. Both were verified against the loan's own payments.
#   row 9  (S No 8)  cell read 06-07-2024, actually 06-03-2024 — April payment then fits +1
#   row 97 (S No 96) cell read 21-11-2021, actually 21-11-2024 — Dec 2024 payment fits +1
ROW_DATE_OVERRIDES = {
    9:  date(2024, 3, 6),
    97: date(2024, 11, 21),
}

MARKUP = 1.25
SERVICE_CHARGE = 0.01

# The month headers were text like "Mar-24" that Excel misparsed into dates with a
# meaningless year. Only the day component survived, and it carries the month number:
# 3..12 for 2024, 1..12 for 2025, 1..7 for 2026 — 29 columns, matching G..AI exactly.
def month_sequence() -> list[date]:
    months = [(2024, m) for m in range(3, 13)]
    months += [(2025, m) for m in range(1, 13)]
    months += [(2026, m) for m in range(1, 8)]
    return [date(y, m, 1) for y, m in months]


def normalise_phone(value) -> tuple[str | None, str | None]:
    """Returns (valid_phone, raw_if_invalid)."""
    if value in (None, ''):
        return None, None
    raw = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 12 and digits.startswith('91'):
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith('0'):
        digits = digits[1:]
    if re.fullmatch(r'[6-9]\d{9}', digits):
        return digits, None
    return None, raw.strip()


def parse_date(value) -> tuple[date | None, str | None]:
    """
    Column B mixes real date cells with text. 175 of 429 rows store the date as a
    `dd-mm-yyyy` string, and treating only true date cells as valid wrongly reported
    41% of the ledger as undated.

    Returns (parsed, ambiguity_note). Indian dd-mm-yyyy is assumed; a note is raised
    when both leading parts are <= 12 and the order cannot be proven from the value.
    """
    if value in (None, ''):
        return None, None
    if hasattr(value, 'year'):
        # Excel date cells in this sheet are corrupt: the meaningful month ended up in
        # the DAY slot while the year survived. Proven on the header row, where this rule
        # recovers all 29 month labels exactly (Mar 2024 .. Jul 2026), and confirmed by
        # behaviour — the 159 rows stored as text need no correction and 88% of them pay
        # one month after disbursement, whereas the 240 date cells produce 161 payments
        # dated before their own loan until this is applied.
        #
        # The day-of-month is unrecoverable, so it is set to 1 and the loan is flagged.
        cell = value.date() if hasattr(value, 'date') else value
        month = cell.day
        if 1 <= month <= 12:
            return date(cell.year, month, 1), f'month-precision only (recovered from {cell.isoformat()})'
        return cell, f'date cell outside recoverable range: {cell.isoformat()}'

    text = str(value).strip()
    if text in DATE_CORRECTIONS:
        text = DATE_CORRECTIONS[text]
    m = re.fullmatch(r'(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})', text)
    if m:
        first, second, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        # A value above 12 in either slot settles the order outright.
        if first > 12:
            day, month = first, second
        elif second > 12:
            day, month = second, first
        else:
            day, month = first, second       # dd-mm, consistent with the unambiguous rows
            try:
                return date(year, month, day), f'ambiguous day/month: {text}'
            except ValueError:
                return None, f'unparseable: {text}'
        try:
            return date(year, month, day), None
        except ValueError:
            return None, f'out of range: {text}'

    for fmt in ('%d-%b-%Y', '%d %b %Y', '%d-%B-%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date(), None
        except ValueError:
            pass
    return None, f'unrecognised date: {text}'


def money(value) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 2)
    return None


def read_images(xlsx: Path) -> dict[int, dict[str, str]]:
    """row index (1-based sheet row) -> {'photo': file, 'aadhaar': file}"""
    z = zipfile.ZipFile(xlsx)
    drawing = z.read('xl/drawings/drawing1.xml').decode('utf-8', 'replace')
    rels_raw = z.read('xl/drawings/_rels/drawing1.xml.rels').decode('utf-8', 'replace')
    rels = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_raw))

    # Each oneCellAnchor holds a from-position and one embedded image.
    anchors = re.findall(
        r'<xdr:oneCellAnchor>.*?<xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>'
        r'.*?r:embed="(rId\d+)".*?</xdr:oneCellAnchor>',
        drawing, re.S)

    PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    by_row: dict[int, dict[str, str]] = defaultdict(dict)
    for col, row, rid in anchors:
        target = rels.get(rid)
        if not target:
            continue
        col_i, row_i = int(col), int(row)
        kind = 'photo' if col_i == COL_PHOTO - 1 else 'aadhaar' if col_i == COL_AADHAAR - 1 else None
        if kind is None:
            continue
        member = 'xl/' + target.replace('../', '')
        sheet_row = row_i + 1
        name = f"row{sheet_row:04d}_{kind}{Path(member).suffix}"
        with z.open(member) as src, open(PHOTO_DIR / name, 'wb') as dst:
            shutil.copyfileobj(src, dst)
        by_row[sheet_row][kind] = name
    return by_row


def read_hyperlinks(xlsx: Path) -> dict[int, str]:
    """
    row -> location URL from the Location column.

    Read from the raw XML rather than openpyxl: the attributes appear as
    `r:id` then `ref`, and anything assuming the opposite order silently finds
    nothing — which is exactly what happened on the first pass.
    """
    z = zipfile.ZipFile(xlsx)
    sheet = z.read('xl/worksheets/sheet1.xml').decode('utf-8', 'replace')
    rels = dict(re.findall(
        r'Id="(rId\d+)"[^>]*Target="([^"]+)"',
        z.read('xl/worksheets/_rels/sheet1.xml.rels').decode('utf-8', 'replace')))

    out: dict[int, str] = {}
    for tag in re.findall(r'<hyperlink\b[^>]*/?>', sheet):
        rid = re.search(r'r:id="(rId\d+)"', tag)
        ref = re.search(r'ref="([A-Z]+)(\d+)"', tag)
        if not (rid and ref):
            continue
        if openpyxl.utils.column_index_from_string(ref.group(1)) != COL_LOCATION:
            continue
        target = rels.get(rid.group(1))
        if target:
            out[int(ref.group(2))] = target
    return out


def pick_name(variants: list[tuple[str, date | None]]) -> tuple[str, list[str]]:
    """
    Most frequent spelling wins; ties break on the longer (usually more complete)
    form, then on the most recent loan. Returns (chosen, discarded).
    """
    counts = Counter(n for n, _ in variants)
    latest = {}
    for n, d in variants:
        if d and (n not in latest or d > latest[n]):
            latest[n] = d
    ranked = sorted(
        counts,
        key=lambda n: (counts[n], len(n), latest.get(n) or date.min),
        reverse=True,
    )
    return ranked[0], ranked[1:]


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("usage: extract-ledger.py <path to link.xlsx>")
    xlsx = Path(sys.argv[1])
    OUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[SHEET]
    months = month_sequence()
    images = read_images(xlsx)
    links = read_hyperlinks(xlsx)

    rows, date_warnings = [], []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        get = lambda c: r[c - 1].value  # noqa: E731
        sno, name, phone_raw = get(COL_SNO), get(COL_NAME), get(COL_PHONE)
        if all(v in (None, '') for v in (sno, name, phone_raw)):
            continue

        phone, phone_invalid = normalise_phone(phone_raw)
        sheet_row_no = r[0].row
        if sheet_row_no in ROW_DATE_OVERRIDES:
            loan_date, date_note = ROW_DATE_OVERRIDES[sheet_row_no], None
        else:
            loan_date, date_note = parse_date(get(COL_DATE))
        if date_note:
            date_warnings.append({'sheetRow': r[0].row, 'note': date_note})

        payments = []
        for offset, col in enumerate(range(MONTH_COL_FIRST, MONTH_COL_LAST + 1)):
            amount = money(get(col))
            if amount:
                payments.append({'month': months[offset].isoformat(), 'amount': amount})

        sheet_row = r[0].row
        rows.append({
            'sheetRow': sheet_row,
            'serial': int(sno) if isinstance(sno, (int, float)) else sno,
            'dateGiven': loan_date.isoformat() if loan_date else None,
            'name': str(name).strip() if name else None,
            'phone': phone,
            'phoneInvalid': phone_invalid,
            'amountOut': money(get(COL_OUT)),
            'amountIn': money(get(COL_IN)),
            'area': str(get(COL_AREA)).strip() if get(COL_AREA) else None,
            'surety': str(get(COL_SURETY)).strip() if get(COL_SURETY) else None,
            'locationUrl': links.get(sheet_row),
            'photo': images.get(sheet_row, {}).get('photo'),
            'aadhaar': images.get(sheet_row, {}).get('aadhaar'),
            'payments': payments,
            'totalPaid': round(sum(p['amount'] for p in payments), 2),
            '_dateObj': loan_date,
        })

    # ---- borrowers: phone is the identity. Rows without one become their own borrower,
    # keyed on sheet row so two unrelated people are never merged by a shared name.
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        key = f"phone:{row['phone']}" if row['phone'] else f"row:{row['sheetRow']}"
        groups[key].append(row)

    borrowers, warnings = [], []
    for key, group in groups.items():
        variants = [(g['name'], g['_dateObj']) for g in group if g['name']]
        chosen, discarded = pick_name(variants) if variants else ('(unnamed)', [])
        if discarded:
            warnings.append({
                'type': 'name_variants_merged',
                'phone': group[0]['phone'],
                'kept': chosen,
                'discarded': discarded,
                'loans': len(group),
            })
        picked = next((g for g in group if g['photo'] or g['aadhaar'] or g['locationUrl']), group[0])
        borrowers.append({
            'key': key,
            'name': chosen,
            'mobile': group[0]['phone'],
            'mobileMissing': group[0]['phone'] is None,
            'mobileInvalidRaw': next((g['phoneInvalid'] for g in group if g['phoneInvalid']), None),
            'area': next((g['area'] for g in group if g['area']), None),
            'locationUrl': picked['locationUrl'],
            'photo': picked['photo'],
            'aadhaar': picked['aadhaar'],
            'loanSerials': [g['serial'] for g in group],
            'loanCount': len(group),
        })

    # ---- loans: derive the schema's figures. IN is the reliable field; OUT is used as
    # the amount actually handed over when present, otherwise derived from the 1% charge.
    seen_serials: Counter = Counter()
    loans = []
    for row in rows:
        total_repayment = row['amountIn']
        principal = round(total_repayment / MARKUP, 2) if total_repayment else None
        if principal:
            derived_out = round(principal * (1 - SERVICE_CHARGE), 2)
            received = row['amountOut'] if row['amountOut'] else derived_out
            if row['amountOut'] and abs(row['amountOut'] - derived_out) > 1:
                warnings.append({
                    'type': 'out_does_not_match_1pc_charge',
                    'serial': row['serial'], 'sheetRow': row['sheetRow'],
                    'amountOut': row['amountOut'], 'expected': derived_out,
                })
        else:
            received = row['amountOut']
            warnings.append({
                'type': 'missing_amount_in',
                'serial': row['serial'], 'sheetRow': row['sheetRow'],
            })

        serial = row['serial']
        seen_serials[serial] += 1
        suffix = seen_serials[serial]
        display_serial = str(serial) if suffix == 1 else f"{serial}-{chr(ord('a') + suffix - 2)}"
        if suffix > 1:
            warnings.append({
                'type': 'duplicate_serial',
                'serial': serial, 'sheetRow': row['sheetRow'], 'renamedTo': display_serial,
            })

        instalments = len(row['payments'])
        loans.append({
            'sheetRow': row['sheetRow'],
            'serial': serial,
            'displaySerial': display_serial,
            'borrowerKey': f"phone:{row['phone']}" if row['phone'] else f"row:{row['sheetRow']}",
            'dateGiven': row['dateGiven'],
            'primaryAmount': principal,
            'amountUserReceived': received,
            'totalRepayment': total_repayment,
            'serviceChargePercent': SERVICE_CHARGE * 100,
            'markupPercent': (MARKUP - 1) * 100,
            'payments': row['payments'],
            'instalmentsPaid': instalments,
            'totalPaid': row['totalPaid'],
            'outstanding': round((total_repayment or 0) - row['totalPaid'], 2),
        })

    for dw in date_warnings:
        warnings.append({'type': 'date_needs_review', **dw})

    payload = {
        'source': str(xlsx.name),
        'sheet': SHEET,
        'monthColumns': [m.isoformat() for m in months],
        'borrowers': borrowers,
        'loans': loans,
        'warnings': warnings,
    }
    (OUT_DIR / 'ledger.json').write_text(json.dumps(payload, indent=2, ensure_ascii=False))

    # ---- summary (counts only; no names, numbers or photos printed)
    wc = Counter(w['type'] for w in warnings)
    photos = sum(1 for b in borrowers if b['photo'])
    aadhaars = sum(1 for b in borrowers if b['aadhaar'])
    print(f"loans     : {len(loans)}")
    print(f"borrowers : {len(borrowers)}  "
          f"(missing mobile {sum(1 for b in borrowers if b['mobileMissing'])}, "
          f"multi-loan {sum(1 for b in borrowers if b['loanCount'] > 1)})")
    print(f"payments  : {sum(len(l['payments']) for l in loans)} cells")
    print(f"photos    : {photos} profile, {aadhaars} aadhaar -> {PHOTO_DIR}/")
    print(f"locations : {sum(1 for b in borrowers if b['locationUrl'])}")
    print("\nwarnings:")
    for t, n in wc.most_common():
        print(f"  {n:>4}  {t}")
    print(f"\nwritten: {OUT_DIR / 'ledger.json'}")


if __name__ == '__main__':
    main()
